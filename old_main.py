from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import TimeoutException
from selenium import webdriver
import time
import requests
from bs4 import BeautifulSoup
import datetime
import schedule
import openpyxl
from openpyxl.styles import PatternFill
from old_is_it_sold_NOTFIXED import check_sold_ads
import re


def get_ads(url):
    for _ in range(2):  # Try up to 2 times
        options = Options()
        options.add_argument("--headless")  # Run Chrome in headless mode (without GUI)

        # Automatically download and set up ChromeDriver
        driver = webdriver.Chrome(options=options)

        # Set the timeout
        driver.set_page_load_timeout(10)

        try:
                driver.get(url)
                time.sleep(2)

                page_source = driver.page_source

                # Parse the HTML content of the page with BeautifulSoup
                soup = BeautifulSoup(page_source, 'html.parser')

                # Find all ads - now looking for 'a' elements with class 'sf-search-ad-link link link--dark hover:no-underline'
                ads = soup.find_all('a', class_='sf-search-ad-link link link--dark hover:no-underline')

                # Prepare a list to hold the results
                results = []

                # Iterate over the first 6 ads found
                for ad in ads[:6]:
                    # If a hyperlink was found in the ad, append it to the results
                    if ad and ad.has_attr('href'):
                        results.append(ad['href'])

                # Return the results
                driver.quit()
                return results

        except TimeoutException:
                print("Page took too long to load. Quitting.")
                driver.quit()


def get_ad_info(url):
    for _ in range(2):  # Try up to 2 times
        try:
            r = requests.get(url)
            r.raise_for_status()

            soup = BeautifulSoup(r.text, 'html.parser')

            title = soup.find('h1', class_='u-t2 u-word-break').text
            title_words = title.split(' ')
            make = title_words[0]
            model = ' '.join(title_words[1:])  # Join the rest of the words back into a single string

            year = int(soup.find('div', class_='u-strong').text)
  
            km = int(soup.find_all('div', class_='u-strong')[1].text.replace('\xa0', '').replace(' km', ''))
            km = round(km/10000) * 10000
            # If km is a 7-digit number, remove the last 0
            if 1000000 <= km < 10000000:
                km = km // 10

            # Find the dd element that follows dt with 'Reg.nr.'
            reg_nr = None
            for dt in soup.find_all('dt'):
                if dt.text.strip() == 'Reg.nr.':
                    reg_nr = dt.find_next_sibling('dd').text
                    break

            date = datetime.date.today().strftime('%d.%m.%Y')

            # Find the dd element that follows dt with 'Pris eks omreg'
            price = None
            for dt in soup.find_all('dt'):
                if dt.text.strip() == 'Pris eks omreg':
                    price = int(dt.find_next_sibling('dd').text.replace('\xa0', '').replace(' kr', ''))
                    # Round price to the nearest 1000
                    price = round(price / 1000) * 1000
                    break

            # Get the description
            description_h2 = soup.find('h2', class_='u-t3', string='Beskrivelse')
            description_p = description_h2.find_next_sibling('p') if description_h2 else None
            description_text = description_p.text if description_p else ''

            # Convert the description and search words to lowercase for case-insensitive search
            description = description_text.lower()
            words_delebil = ["delebil", "rep objekt", "rep. objekt", "repobjekt", "reperasjonsobjekt", "reparasjons objekt", "dele bil", "ikke eu", "mangler eu",
                     "ikke godkjent", "motorhavari", "feil med motor", "defekt motor", "bilen er ikke kjørbar"]

            # Initialize the description rating to None
            description_rating = None

            # Check if any of the words exist in the description
            for word in words_delebil:
                if word in description:
                    description_rating = "Delebil"
                    break

            kr = "kr"

            span = soup.find('span', class_='u-mh16', string=re.compile(r'^\d{4}\s'))
            if span:
                postal_code = span.text.split()[0]
            else:
                postal_code = None

            # Determine place based on the first digit of postal_code
            if postal_code:
                first_digit = postal_code[0]
                if first_digit == '0':
                    place = "Oslo"
                elif first_digit == '1':
                    place = "Østfold"
                elif first_digit == '2':
                    place = "Oppland"
                elif first_digit == '3':
                    place = "Vestfold"
                elif first_digit == '4':
                    place = "Sørnorge"
                elif first_digit == '5':
                    place = "Hordaland"
                elif first_digit == '6':
                    place = "Sogn og Fjordane"
                elif first_digit == '7':
                    place = "Trønderlag"
                elif first_digit == '8':
                    place = "Nordalnd"
                elif first_digit == '9':
                    place = "Til helvete"
                else:
                    place = "Unknown"
            else:
                place = "Unknown"

            return {
                'Make': make,
                'Model': model,
                'Description': description_rating,  # Add the description rating to the returned dictionary
                'Reg_nr': reg_nr,
                'Year': year,
                'KM': km,
                'Place': place,
                'Price': price,
                'Kr': kr,
                'Date': date,
                'URL': url,
            }

        except (requests.RequestException, ValueError):
            print(f"Failed to fetch {url}, retrying in 5 seconds...")
            time.sleep(5)  # Wait for 5 seconds before trying again


def add_row_to_sheet(ws, ad_info):
    # Print for å se hva som blir lagt inn
    print("Legges inn i filen \n")
    # Remove 'URL' from ad_info
    ad_info_without_url = {k: v for k, v in ad_info.items() if k not in ['URL']}
    row_values = list(ad_info_without_url.values())

    last_row = max((c.row for c in ws['A'] if c.value is not None))  # assuming column 'A' should always have data

    for i, value in enumerate(row_values, start=1):
        cell = ws.cell(row=last_row + 1, column=i, value=value)
        cell.value = value  # assign the value

        if i in [6, 8]:  # columns F and G
            cell.number_format = '# ##0'  # number with thousands separator

    # Write the URL in column 'P' of the current row
    ws.cell(row=last_row + 1, column=13, value=ad_info['URL']).hyperlink = ad_info['URL']

    similar_ad_found = False
    similar_ad_sold = 0  # Counter for similar ads
    similar_ad_not_sold = 0

    # Check for similar ad
    for row in ws.iter_rows(min_row=2, max_row=last_row):  # iterating from row 2 to last row
        try:
            if ad_info['Model'] is not None:

                if row[1].value == ad_info['Model']:

                    if row[2].value == ad_info['Description']:

                        if row[4].value is not None:

                            if abs(int(row[4].value) - ad_info['Year']) <= 1:

                                if row[5].value is not None:

                                    km_row = int(row[5].value.replace(" ", "")) if isinstance(row[5].value,
                                                                                              str) else row[5].value
                                    km_ad_info = int(ad_info['KM'])

                                    if abs(km_row - km_ad_info) <= 30000:

                                        if ad_info['Price'] is not None and row[10].value is not None:

                                            price_row = int(row[10].value.replace(" ", "")) if isinstance(
                                                row[10].value, str) else row[10].value
                                            price_ad_info = int(
                                                ad_info['Price'])

                                            if price_ad_info <= price_row:
                                                url_row = row[15].value

                                                # Antall dager solgt ikke er None
                                                if row[14].value is not None:

                                                    if int(row[14].value) == 0:

                                                        print(f"Jackpot! Kan bli solgt for {price_row}, "
                                                              f"og koster {price_ad_info}. Den er lik denne: {url_row}")
                                                        similar_ad_found = True
                                                        similar_ad_sold += 1

                                                    elif int(row[14].value) == 1:
                                                        similar_ad_found = True
                                                        similar_ad_sold += 0.5

                                                    else:
                                                        similar_ad_not_sold += 1

        except ValueError:
            # If a ValueError occurs (likely due to attempting to cast a non-integer string), skip this row
            continue
    # If no similar ad is found, highlight the newly added row
    if similar_ad_found:
        print(f"Antall like som ble solgt samme dag: {similar_ad_sold} \n"
              f"Antall like som ikke ble solgt samme dag: {similar_ad_not_sold} \n")

        if similar_ad_not_sold == 0:
            if similar_ad_sold == 1:
                for column in ws.iter_cols(min_col=1, max_col=len(row_values), min_row=last_row + 1, max_row=last_row + 1):
                    for cell in column:
                        cell.fill = yellow_fill_2
            elif similar_ad_sold > 1:
                for column in ws.iter_cols(min_col=1, max_col=len(row_values), min_row=last_row + 1,
                                           max_row=last_row + 1):
                    for cell in column:
                        cell.fill = yellow_fill_1
            elif similar_ad_sold < 1:
                pass
        # Hvis forholdet er under 1/5
        elif similar_ad_not_sold / similar_ad_sold <= 0.20:
            for column in ws.iter_cols(min_col=1, max_col=len(row_values), min_row=last_row + 1, max_row=last_row + 1):
                for cell in column:
                    cell.fill = yellow_fill_1
        # Hvis forholdet er under 1/2
        elif similar_ad_not_sold / similar_ad_sold <= 0.5:
            for column in ws.iter_cols(min_col=1, max_col=len(row_values), min_row=last_row + 1, max_row=last_row + 1):
                for cell in column:
                    cell.fill = yellow_fill_2
        # Hvis forholdet er under 0.7
        elif similar_ad_not_sold / similar_ad_sold <= 0.7:
            for column in ws.iter_cols(min_col=1, max_col=len(row_values), min_row=last_row + 1, max_row=last_row + 1):
                for cell in column:
                    cell.fill = yellow_fill_3

        else:
            print(f"For mange ikke solgt. Forholdet er {similar_ad_not_sold/similar_ad_sold} \n")


def main():
    url = 'https://www.finn.no/car/used/search.html?d' \
          'ealer_segment=3&price_to=200000&published=1&sales_form=1&sort' \
          '=PUBLISHED_DESC&stored-id=60509806&year_from=1990'

    ads = get_ads(url)
    print("\n", ads)
    # Get the current time
    current_time = datetime.datetime.now()

    # Print the current time
    print("Current time:", current_time)

    # Open the Excel Workbook
    wb = openpyxl.load_workbook('data.xlsx')
    time.sleep(1)

    # Check if ads is not None
    if ads is not None:

        # Loop through all the advertisements
        for ad in ads:
            ad_info = get_ad_info(ad)
            if ad_info is None:
                print(f"Failed to get info for ad: {ad}")
                continue

            # Print only the specified fields from the information for each ad
            print(f"Make: {ad_info.get('Make')} {ad_info.get('Model')} i {ad_info.get('Place')} "
                  f"\nLink: {ad_info.get('URL')}")

            # Determine the sheet to write to based on 'Make'
            make = ad_info.get('Make')
            if make in wb.sheetnames:
                ws = wb[make]
            else:
                # If the sheet doesn't exist, use 'Others' sheet. Create it if it doesn't exist.
                if 'Others' not in wb.sheetnames:
                    wb.create_sheet('Others')
                ws = wb['Others']

            # Check if the reg_nr of the current ad is already in the sheet

            # Get the total number of rows in the sheet
            total_rows = ws.max_row
            # Calculate the starting row for the last 200 rows
            start_row = max(1, total_rows - 200 + 1)
            # Create a list from the last 200 cells (or fewer if there aren't 200 rows) in column 'D'
            last_200_cells = [cell.value for cell in ws['D'][start_row - 1:] if cell.value]
            # Convert the list to a set to remove duplicates and speed up 'in' checks
            existing_reg_nrs = set(last_200_cells)

            if ad_info['Reg_nr'] is not None and ad_info['Reg_nr'] not in existing_reg_nrs:
                # Add the row to the appropriate sheet
                add_row_to_sheet(ws, ad_info)

    else:
        print("No ads to process.")

    # Save workbook after adding all the new entries
    wb.save('data.xlsx') 


main()

schedule.every().day.at("23:20").do(check_sold_ads)

schedule.every(100).seconds.do(main)

while True:
    schedule.run_pending()
    time.sleep(1)
