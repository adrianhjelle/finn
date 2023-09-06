import openpyxl
import requests
from bs4 import BeautifulSoup
import datetime
import validators


def get_page(url):
    try:
        r = requests.get(url)
        r.raise_for_status()
        return r.text
    except requests.exceptions.HTTPError as err:
        if err.response.status_code == 410:
            print(f"The URL {url} is no longer available.")
            return "Fjernet"
        else:
            raise


def check_sold_ads():
    # Open the workbook
    workbook = openpyxl.load_workbook('data.xlsx')

    # Iterate over all sheets in the workbook
    for sheet in workbook.worksheets:

        print(f"Processing sheet: {sheet.title}")

        # Iterate over the rows in the current sheet
        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            url = row[15]  # URL is in the 12 column
            current_date = datetime.date.today()

            if row[13] is not None or row[14] is not None:
                continue

            try:
                if row[12]:  # Check if the date in column I has something in it
                    if isinstance(row[12], datetime.datetime):  # Check if it's already a datetime object
                        date_in_column_i = row[12].date()
                    else:
                        date_in_column_i = datetime.datetime.strptime(row[12], '%d.%m.%Y').date()

                    date_diff = (current_date - date_in_column_i).days

                    if date_diff >= 6:
                        cell = sheet.cell(row=index, column=15)  # get the cell in column K
                        cell.value = "x"  # set the value of the cell
                        continue

                else:
                    continue

            except ValueError:
                print(f"Row {index}: The date in column I is not in the 'dd.mm.yyyy' format.")
                continue

            # Skip to the next row if URL is not valid
            if url is None or not validators.url(url):
                continue

            try:
                page = get_page(url)
                current_date = datetime.date.today()
                if page == "Fjernet":
                    # Write the current date in column J (index 9) if the ad is sold
                    cell = sheet.cell(row=index, column=14)
                    cell.value = current_date.strftime('%d.%m.%Y')

                    # Calculate the difference between current date and date in column I
                    if row[12]:
                        cell = sheet.cell(row=index, column=15)
                        cell.value = date_diff  # Writing the date difference
                    continue

                if page is not None:
                    soup = BeautifulSoup(page, 'html.parser')

                    sold_status = soup.find('span', {'class': 'u-capitalize status status--warning u-mb0'})

                    if sold_status and sold_status.text.strip().lower() == 'solgt':
                        # Write the current date in column J (index 9) if the ad is sold
                        cell = sheet.cell(row=index, column=14)
                        cell.value = current_date.strftime('%d.%m.%Y')

                        # Calculate the difference between current date and date in column I
                        if row[12]:
                            cell = sheet.cell(row=index, column=15)
                            cell.value = date_diff  # Writing the date difference

                    elif sold_status and sold_status.text.strip().lower() == 'inaktiv':
                        # Write the current date in column J (index 9) if the ad is inaktiv
                        cell = sheet.cell(row=index, column=14)
                        cell.value = current_date.strftime('%d.%m.%Y')

                        # Calculate the difference between current date and date in column I
                        if row[8]:
                            cell = sheet.cell(row=index, column=15)
                            cell.value = "Inaktiv"

                else:
                    print("Unable to fetch page content.")  # Notify if unable to fetch the page content
                    continue  # Skip this iteration and move to the next row/url

            except Exception as e:
                print(f"An error occurred while processing the URL {url}: {e}")
                continue  # Skip this iteration and move to the next row/url

        # Save the workbook
        workbook.save('data.xlsx')

    print("Da var alle linkene sjekket!")


if __name__ == "__main__":
    check_sold_ads()
